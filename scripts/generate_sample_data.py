"""Generate realistic-looking LinkedIn export workbooks for local testing.

Usage (from repo root):
    python scripts/generate_sample_data.py

Writes three .xlsx files into ./sample_data that can be dropped into the
dashboard's upload page. Header rows are intentionally offset to exercise the
header-detection heuristic.
"""
from __future__ import annotations

import os
import random
from datetime import date, timedelta

import openpyxl

random.seed(42)
OUT = os.path.join(os.path.dirname(__file__), "..", "sample_data")
os.makedirs(OUT, exist_ok=True)

START = date(2024, 1, 1)
DAYS = 120


def _save(wb: openpyxl.Workbook, name: str) -> None:
    path = os.path.join(OUT, name)
    wb.save(path)
    print(f"wrote {path}")


def followers() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "New followers"
    ws.append(["Follower analytics", None])           # metadata row 1
    ws.append(["Generated for demo", None])           # metadata row 2
    ws.append(["Date", "Total followers", "Organic followers", "Sponsored followers"])
    base = 20
    for i in range(DAYS):
        d = START + timedelta(days=i)
        organic = max(0, int(random.gauss(base + i * 0.3, 8)))
        sponsored = random.randint(0, 5)
        ws.append([d.strftime("%m/%d/%Y"), organic + sponsored, organic, sponsored])

    # Demographic tabs.
    demos = {
        "Job function": ["Engineering", "Sales", "Marketing", "Operations", "Finance"],
        "Seniority": ["Senior", "Manager", "Director", "Entry", "VP"],
        "Industry": ["Software", "Finance", "Healthcare", "Retail", "Education"],
        "Location": ["United States", "India", "United Kingdom", "Germany", "Brazil"],
        "Company size": ["1-10", "11-50", "51-200", "201-1000", "1000+"],
    }
    for dim, cats in demos.items():
        sheet = wb.create_sheet(dim[:31])
        sheet.append([dim, "Total followers"])
        for c in cats:
            sheet.append([c, random.randint(80, 1200)])
    _save(wb, "followers.xlsx")


def visitors() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Visitors"
    ws.append(["Visitor highlights", None])
    ws.append(["Date", "Total page views", "Total unique visitors", "Desktop page views", "Mobile page views"])
    for i in range(DAYS):
        d = START + timedelta(days=i)
        uv = max(10, int(random.gauss(120 + i, 25)))
        pv = int(uv * random.uniform(1.4, 2.0))
        desktop = int(pv * random.uniform(0.5, 0.7))
        ws.append([d.strftime("%m/%d/%Y"), pv, uv, desktop, pv - desktop])
    _save(wb, "visitors.xlsx")


def content() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All posts"
    ws.append(["Post analytics export", None])
    ws.append(["Post URL", "Post title", "Post type", "Created date", "Impressions",
               "Clicks", "Reactions", "Comments", "Reposts"])
    types = ["video", "image", "article", "text", "document"]
    for i in range(45):
        d = START + timedelta(days=random.randint(0, DAYS - 1))
        ptype = random.choice(types)
        imp = random.randint(2000, 60000)
        # videos engage harder; text reaches less.
        eng_mult = {"video": 1.8, "document": 1.4, "image": 1.0, "article": 0.9, "text": 0.6}[ptype]
        reactions = int(imp * 0.02 * eng_mult * random.uniform(0.7, 1.3))
        comments = int(reactions * 0.1)
        reposts = int(reactions * 0.05)
        clicks = int(imp * random.uniform(0.002, 0.02))
        ws.append([
            f"https://linkedin.com/feed/update/{i}", f"Sample post #{i}", ptype,
            d.strftime("%m/%d/%Y"), imp, clicks, reactions, comments, reposts,
        ])
    _save(wb, "content.xlsx")


if __name__ == "__main__":
    followers()
    visitors()
    content()
    print("Done. Drop the files in ./sample_data into the Uploads page.")
